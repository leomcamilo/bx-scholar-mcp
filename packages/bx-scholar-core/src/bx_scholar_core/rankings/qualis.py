"""Qualis CAPES rankings loader — XLSX via openpyxl, no pandas."""

from __future__ import annotations

from pathlib import Path

from bx_scholar_core.logging import get_logger
from bx_scholar_core.models.ranking import QualisEntry

logger = get_logger(__name__)


def load_qualis(path: Path) -> dict[str, QualisEntry]:
    """Load Qualis CAPES rankings from XLSX.

    Returns issn -> QualisEntry mapping.
    """
    index: dict[str, QualisEntry] = {}

    if not path.exists():
        logger.warning("qualis_not_found", path=str(path))
        return index

    try:
        import openpyxl

        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        if ws is None:
            logger.warning("qualis_no_active_sheet")
            return index

        headers = [str(cell.value or "").lower() for cell in next(ws.iter_rows(max_row=1))]

        # Find column indices
        issn_col = next((i for i, h in enumerate(headers) if "issn" in h), None)
        if issn_col is None:
            logger.warning("qualis_no_issn_column", headers=headers)
            return index

        title_col = next((i for i, h in enumerate(headers) if "tulo" in h or "title" in h), None)
        qualis_col = next(
            (i for i, h in enumerate(headers) if "estrato" in h or "qualis" in h), None
        )
        area_col = next((i for i, h in enumerate(headers) if "rea" in h and len(h) > 2), None)

        for row in ws.iter_rows(min_row=2, values_only=True):
            if issn_col >= len(row):
                continue
            issn = str(row[issn_col] or "").strip()
            if not issn or issn == "nan":
                continue

            entry = QualisEntry(
                title=str(row[title_col] or "")
                if title_col is not None and title_col < len(row)
                else "",
                classification=str(row[qualis_col] or "")
                if qualis_col is not None and qualis_col < len(row)
                else "",
                area=str(row[area_col] or "")
                if area_col is not None and area_col < len(row)
                else "",
            )
            index[issn] = entry

        wb.close()
        logger.info("qualis_loaded", count=len(index))
    except Exception as exc:
        logger.error("qualis_load_failed", error=str(exc))

    return index
